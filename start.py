# coding=gbk
#from xlrd import xlsx
from openpyxl import load_workbook
import datetime
import os

STARTING_LINE_ROW_OFFSET = 8
INDEX_OFFSET = 0
PICTURE_OFFSET = INDEX_OFFSET + 1
NAME_OFFSET = PICTURE_OFFSET + 1
UNIT_OFFSET = NAME_OFFSET + 1
SEND_OFFSET = UNIT_OFFSET + 1
RECV_OFFSET = SEND_OFFSET + 1
COMMENT_OFFSET = RECV_OFFSET + 1

START_NEW_LINE_STR= "[��������]"
QUANTITY= "[����]"

NAME_AND_OFFSETS = [
    ("[��������]", 'index', INDEX_OFFSET),
    ("[��Ʒ�ͺ�]", 'picture', PICTURE_OFFSET),
    ("[��Ʒ����]", 'name', NAME_OFFSET),
    ("[����]", 'quantity_send', SEND_OFFSET),
    ("[����]", 'quantity_recv', RECV_OFFSET),    
    ("[�ͻ��ɹ�����]", 'comments', COMMENT_OFFSET),    
]
  
 
def read_src_file(filename):    
    try:
        fi = open(filename)
    except:
        print("{}{}{}".format("�ļ���", filename, " ������"))
        return None
    lines = fi.readlines()

    context = []    
    item = {}
    for line in lines:
        for _ in NAME_AND_OFFSETS:
            if _[0] in line:
                if START_NEW_LINE_STR == _[0]:
                    if item:                       
                        context.append(item)                    
                    item = {}
                    try:
                        value = line.split()[0]
                        #print(value)
                        item[_[2]] = value
                    except:
                        print("{}{}".format(line.strip('\n'), " ��ʽ����"))
                    continue
                try:
                    value = line.split(":")[1]
                    item[_[2]] = value
                except:
                    print("{}{}".format(line.strip('\n'), " ��ʽ����"))   
                    
                if QUANTITY == _[0]:
                    item[RECV_OFFSET] = value                    
    if item:
        context.append(item)
        
    #print(context)
    
    fi.close()
    
    return context


context = read_src_file('import.txt')

if not context:
    print("���ɱ���ʧ��!")
    exit()

try:
    wb = load_workbook("myopenpyxl.xlsx")
except:
    print("����ģ��ʧ��!")
    exit()

ws = wb[wb.sheetnames[0]]

for i,line in enumerate(context):
    for key, value in line.items():
        if key == len(line.items()) and i:
            pass
        else:
            ws.cell(row=STARTING_LINE_ROW_OFFSET + i, column=key+1, value=value)
    ws.cell(row=STARTING_LINE_ROW_OFFSET + i, column=UNIT_OFFSET+1, value="PIC")

timestamp = datetime.datetime.now().strftime('%Y%m%d%H%M%S')

try:
    wb.save("output-{}.xlsx".format(timestamp)) # timestamp
except:
    print("�����ļ�ʧ��")
    exit()








